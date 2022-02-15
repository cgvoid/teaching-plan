from openpyxl import load_workbook

class Anode: 
    def __init__(self, adjvex, weight=0):
        self.Adjvex = adjvex 
        self.Weight = weight


class Vnode: 
    def __init__(self, data):
        self.Data = data 
        self.Firstarc = []


class Graph:
    def __init__(self):
        self.vertList = []
        self.numVertics = 0

    def add_vertex(self, key):
        vertex = Vnode(key)
        self.vertList.append(vertex)
        self.numVertics = self.numVertics + 1
        return vertex

    def add_edge(self, val1, val2, weight=0):
        i = 0
        while i < len(self.vertList):  
            if val1 == self.vertList[i].Data:
                vnode1 = self.vertList[i]
                break
            i = i + 1
        if i == len(self.vertList):  
            vnode1 = self.add_vertex(val1)

        i = 0
        while i < len(self.vertList):  
            if val2 == self.vertList[i].Data:
                vnode2 = self.vertList[i]
                break
            i = i + 1
        if i == len(self.vertList):  
            vnode2 = self.add_vertex(val2)

        v2id = self.vertList.index(vnode2)
        p = Anode(v2id, weight)

        vnode1.Firstarc.append(p) 

class Stack():
    def __init__(self):
        self.slist = []

    def is_empty(self):
        if self.slist == []:
            return 1
        else:
            return 0

    def pushstack(self, data):
        self.slist.append(data)

    def popstack(self):
        return self.slist.pop()



def searchGraph(graph, cur_vertx_indx): 
    orderSeq = []
    visited_li = [0] * graph.numVertics
    s = Stack()
    #orderSeq.append(graph.vertList[cur_vertx_indx].Data)
    s.pushstack(cur_vertx_indx)
    visited_li[cur_vertx_indx] = 1

    while s.is_empty() !=1:
    	cur_vertx_indx = s.popstack()
    	for node in graph.vertList[cur_vertx_indx].Firstarc:
    		if visited_li[node.Adjvex] == 0:
    			s.pushstack(cur_vertx_indx)
    			s.pushstack(node.Adjvex) 
    			visited_li[node.Adjvex] = 1   
    			orderSeq.append(graph.vertList[node.Adjvex].Data)			 
    			break

    for index in range(graph.numVertics):
    	if visited_li[index] == 0:
    		visited_li[index] = 1   
    		orderSeq.append(graph.vertList[index].Data)

    return orderSeq


if __name__ == '__main__':
	courseGraph = Graph()
	'''
	input: course_graph, i.e., triple in sheet 2 of Metabook.xlsx, such as: 
	(Intro to Computer Processing and C++, Subtopic_in, COMP1011)

	'''
	wb = load_workbook("Metabook.xlsx", read_only = True)
	coursSheet = wb["KG2_courses"]
	for i in range(3,53): # COMP 1011
#	for i in range(54,67): # COMP 1433
#	for i in range(68,93): # COMP 1901
#	for i in range(94,148): # COMP 2021
#	for i in range(149,205): # COMP 2022
#	for i in range(206,233): # COMP 2011	
#	for i in range(234,240): # COMP 2411？
#	for i in range(241,259): # COMP 3133
#	for i in range(260,331): # COMP 3423
#	for i in range(332,375): # COMP 4122
#	for i in range(376,394): # COMP 4133
#	for i in range(395,402): # COMP 4422
#	for i in range(403,419): # COMP 4431
#	for i in range(420,428): # COMP 4434
#	for i in range(429,467): # COMP 4334 
		v1 = coursSheet.cell(row = i, column = 5).value
		v2 = coursSheet.cell(row = i, column = 3).value
		courseGraph.add_edge(v1, v2)
	
	startNode = 0 # CourseName do not add in Seq
	Seq = searchGraph(courseGraph, startNode)
	
	'''
	output: sequence of teaching plan, such as: ['C++', 'Control flow', ...]
	
	'''
	print(Seq)
